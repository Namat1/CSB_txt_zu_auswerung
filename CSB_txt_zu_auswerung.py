# CSB_txt_zu_auswerung.py
# Streamlit-App:
# 1) CSB Tour-/Ladeplan TXT auslesen
# 2) Excel-Vorgabe hochladen
# 3) Erste zwei Excel-Blätter mit Kunden/Liefertagen gegen CSB vergleichen
#
# Die Excel ist maßgeblich. Abweichungen zeigen, was in CSB angepasst werden muss.

from __future__ import annotations

import io
import re
import traceback
import zipfile
import xml.etree.ElementTree as ET
from typing import Dict, List, Tuple

import pandas as pd
import streamlit as st


st.set_page_config(
    page_title="CSB Ladeplan Abgleich",
    page_icon="🚚",
    layout="wide",
)

st.title("🚚 CSB Ladeplan Abgleich")
st.success("App ist gestartet.")
st.caption(
    "TXT-Ladeplan und Excel-Vorgabe hochladen. "
    "Die Excel ist maßgeblich und zeigt, welchen Stand CSB annehmen muss."
)
st.info("CSB mit 103/F8 die ganzen Wochentage von 1001-1886 bis 6001-6886 generieren und als TXT exportieren, dann mit Quelldatei abgleichen.")

WOCHENTAG_MAP = {
    "montag": "Mo",
    "dienstag": "Die",
    "mittwoch": "Mitt",
    "donnerstag": "Don",
    "freitag": "Fr",
    "samstag": "Sam",
    "sonntag": "So",
}

TAG_SPALTEN = ["Mo", "Die", "Mitt", "Don", "Fr", "Sam"]


def clean_text(value) -> str:
    if value is None:
        return ""
    value = str(value)
    value = value.replace("\x0c", " ")
    value = value.replace("\xa0", " ")
    value = value.replace("\x81", " ")
    if re.fullmatch(r"\d+\.0", value.strip()):
        value = value.strip()[:-2]
    value = re.sub(r"\s+", " ", value)
    return value.strip(" \t\r\n.;")


def norm_num(value) -> str:
    value = clean_text(value)
    if value == "":
        return ""
    value = value.replace(",", ".")
    if re.fullmatch(r"\d+\.0", value):
        value = value[:-2]
    if re.fullmatch(r"\d+", value):
        return str(int(value))
    return value


def norm_tour(value) -> str:
    return norm_num(value)


def normalize_day(value: str) -> str:
    v = clean_text(value).lower().replace(".", "")
    if v in WOCHENTAG_MAP:
        return WOCHENTAG_MAP[v]
    aliases = {
        "mo": "Mo", "mon": "Mo", "monday": "Mo",
        "die": "Die", "di": "Die", "dienst": "Die", "tuesday": "Die",
        "mitt": "Mitt", "mi": "Mitt", "mittw": "Mitt", "wednesday": "Mitt",
        "don": "Don", "do": "Don", "thursday": "Don",
        "fr": "Fr", "frei": "Fr", "friday": "Fr",
        "sam": "Sam", "sa": "Sam", "saturday": "Sam",
    }
    return aliases.get(v, clean_text(value))


def decode_txt_bytes(data: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("cp1252", errors="replace")


# -------------------------------------------------------------------
# CSB TXT lesen
# -------------------------------------------------------------------

def extract_customer_line(line: str):
    raw = line.rstrip("\r\n").replace("\xa0", " ")

    # Kundenzeilen enden im Ausdruck mit mehreren Punkt-Spalten.
    if not re.search(r"(?:\.\s*){2,}\s*$", raw):
        return None

    # Unterstützt:
    #          10502 Kunde ...
    #     1    13822 Kunde ...
    match_start = re.match(r"^\s{3,}(?:(\d{1,3})\s+)?(\d{3,6})\s+", raw)
    if not match_start:
        return None

    la_aus_txt = match_start.group(1) or ""
    csb = match_start.group(2)

    plz_matches = list(re.finditer(r"\b\d{5}\b", raw))
    if not plz_matches:
        return None

    plz_match = plz_matches[-1]
    plz = plz_match.group(0)

    ort_raw = raw[plz_match.end():]
    ort_raw = re.sub(r"(?:\s+\.){2,}.*$", "", ort_raw)
    ort = clean_text(ort_raw)

    mid = raw[match_start.end():plz_match.start()].rstrip()

    # CSB-Festbreite: Name 21 Zeichen, danach Straße.
    kunde = clean_text(mid[:21])
    strasse = clean_text(mid[21:])

    return la_aus_txt, csb, kunde, strasse, plz, ort


def parse_csb_ladeplan(text: str) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    current_tour = ""
    current_wochentag_raw = ""
    current_liefertag = ""
    current_tour_text = ""
    position = 0

    kunden_rows = []
    tour_meta = {}

    tour_re = re.compile(r"^\s*Tour\s+(\d{3,6})\b(.*?)(?:LKW:|$)", re.IGNORECASE)
    day_re = re.compile(r"^\s*Wochentag\s+(.+?)(?:Fahrer:|$)", re.IGNORECASE)
    count_re = re.compile(r"^\s*(\d+)\s+Anzahl Kunden\b", re.IGNORECASE)

    for raw_line in text.splitlines():
        line = raw_line.rstrip("\n\r")

        day_match = day_re.search(line)
        if day_match:
            current_wochentag_raw = clean_text(day_match.group(1))
            current_liefertag = normalize_day(current_wochentag_raw)

        tour_match = tour_re.search(line)
        if tour_match:
            current_tour = norm_tour(tour_match.group(1))
            current_tour_text = clean_text(tour_match.group(2))
            position = 0

            if not current_liefertag and current_tour:
                current_liefertag = {
                    "1": "Mo", "2": "Die", "3": "Mitt", "4": "Don",
                    "5": "Fr", "6": "Sam", "7": "So",
                }.get(current_tour[0], "")

            tour_meta[current_tour] = {
                "Tour": current_tour,
                "Liefertag": current_liefertag,
                "Wochentag_TXT": current_wochentag_raw,
                "Tour_Text": current_tour_text,
                "Erwartete_Kunden": None,
            }
            continue

        count_match = count_re.search(line)
        if count_match and current_tour:
            tour_meta.setdefault(
                current_tour,
                {
                    "Tour": current_tour,
                    "Liefertag": current_liefertag,
                    "Wochentag_TXT": current_wochentag_raw,
                    "Tour_Text": current_tour_text,
                    "Erwartete_Kunden": None,
                },
            )
            tour_meta[current_tour]["Erwartete_Kunden"] = int(count_match.group(1))
            continue

        customer = extract_customer_line(line)
        if customer and current_tour:
            position += 1
            la_aus_txt, csb, kunde, strasse, plz, ort = customer

            kunden_rows.append(
                {
                    "Quelle": "CSB_TXT",
                    "Tour": current_tour,
                    "Liefertag": current_liefertag,
                    "Wochentag_TXT": current_wochentag_raw,
                    "La_aus_TXT": la_aus_txt,
                    "Position_im_Tourblock": position,
                    "CSB": norm_num(csb),
                    "Kunde": kunde,
                    "Strasse": strasse,
                    "PLZ": norm_num(plz),
                    "Ort": ort,
                    "Tour_Text": current_tour_text,
                }
            )

    kunden_df = pd.DataFrame(kunden_rows)

    if kunden_df.empty:
        empty_touren = pd.DataFrame(
            columns=["Tour", "Liefertag", "Wochentag_TXT", "Tour_Text", "Erwartete_Kunden", "Erkannte_Kunden", "Differenz", "Status"]
        )
        empty_pruefung = pd.DataFrame(
            columns=["Tour", "Liefertag", "Erwartete_Kunden", "Erkannte_Kunden", "Differenz", "Status"]
        )
        return kunden_df, empty_touren, empty_pruefung

    erkannte = (
        kunden_df.groupby("Tour", as_index=False)
        .size()
        .rename(columns={"size": "Erkannte_Kunden"})
    )

    touren_df = pd.DataFrame(tour_meta.values()).merge(erkannte, on="Tour", how="outer")
    touren_df["Erwartete_Kunden"] = pd.to_numeric(touren_df["Erwartete_Kunden"], errors="coerce")
    touren_df["Erkannte_Kunden"] = pd.to_numeric(touren_df["Erkannte_Kunden"], errors="coerce").fillna(0).astype(int)
    touren_df["Differenz"] = touren_df["Erkannte_Kunden"] - touren_df["Erwartete_Kunden"]

    def status(row):
        if pd.isna(row["Erwartete_Kunden"]):
            return "Keine Sollzahl gefunden"
        if row["Differenz"] == 0:
            return "OK"
        return "Abweichung"

    touren_df["Status"] = touren_df.apply(status, axis=1)

    kunden_df = kunden_df.sort_values(["Tour", "Position_im_Tourblock"], kind="stable").reset_index(drop=True)
    touren_df = touren_df.sort_values("Tour", kind="stable").reset_index(drop=True)
    pruefung_df = touren_df[["Tour", "Liefertag", "Erwartete_Kunden", "Erkannte_Kunden", "Differenz", "Status"]].copy()

    return kunden_df, touren_df, pruefung_df


# -------------------------------------------------------------------
# Excel ohne openpyxl lesen: erste zwei Blätter, reine Werte
# -------------------------------------------------------------------

XLSX_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
REL_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"


def excel_col_to_index(cell_ref: str) -> int:
    match = re.match(r"([A-Z]+)", cell_ref)
    if not match:
        return 0
    letters = match.group(1)
    idx = 0
    for ch in letters:
        idx = idx * 26 + (ord(ch) - 64)
    return idx - 1


def load_shared_strings(zf: zipfile.ZipFile) -> List[str]:
    if "xl/sharedStrings.xml" not in zf.namelist():
        return []
    root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    strings = []
    for si in root.findall(XLSX_NS + "si"):
        text_parts = []
        for t in si.iter(XLSX_NS + "t"):
            text_parts.append(t.text or "")
        strings.append("".join(text_parts))
    return strings


def workbook_sheet_targets(zf: zipfile.ZipFile) -> List[Tuple[str, str]]:
    workbook_root = ET.fromstring(zf.read("xl/workbook.xml"))
    rel_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rel_root}

    sheets = []
    sheets_node = workbook_root.find(XLSX_NS + "sheets")
    if sheets_node is None:
        return sheets

    for sheet in sheets_node.findall(XLSX_NS + "sheet"):
        name = sheet.attrib.get("name", "")
        rel_id = sheet.attrib.get(REL_NS + "id")
        target = rel_map.get(rel_id, "")
        if target:
            if target.startswith("/"):
                target = target.lstrip("/")
            elif not target.startswith("xl/"):
                target = "xl/" + target
            sheets.append((name, target))
    return sheets


def read_xlsx_sheet_values(zf: zipfile.ZipFile, target: str, shared_strings: List[str]) -> List[List[str]]:
    root = ET.fromstring(zf.read(target))
    rows = []

    for row_node in root.findall(".//" + XLSX_NS + "sheetData/" + XLSX_NS + "row"):
        row_values: Dict[int, str] = {}
        max_col = -1

        for cell in row_node.findall(XLSX_NS + "c"):
            ref = cell.attrib.get("r", "A1")
            col_idx = excel_col_to_index(ref)
            max_col = max(max_col, col_idx)

            cell_type = cell.attrib.get("t")
            value_node = cell.find(XLSX_NS + "v")
            inline_node = cell.find(XLSX_NS + "is")

            value = ""
            if cell_type == "s" and value_node is not None:
                try:
                    value = shared_strings[int(value_node.text)]
                except Exception:
                    value = ""
            elif cell_type == "inlineStr" and inline_node is not None:
                value = "".join(t.text or "" for t in inline_node.iter(XLSX_NS + "t"))
            elif value_node is not None:
                value = value_node.text or ""

            row_values[col_idx] = clean_text(value)

        if max_col >= 0:
            rows.append([row_values.get(i, "") for i in range(max_col + 1)])

    return rows


def read_excel_first_two_sheets(xlsx_bytes: bytes) -> Dict[str, pd.DataFrame]:
    result = {}

    with zipfile.ZipFile(io.BytesIO(xlsx_bytes)) as zf:
        shared = load_shared_strings(zf)
        sheets = workbook_sheet_targets(zf)[:2]

        for sheet_name, target in sheets:
            rows = read_xlsx_sheet_values(zf, target, shared)

            if not rows:
                result[sheet_name] = pd.DataFrame()
                continue

            max_len = max(len(r) for r in rows)
            rows = [r + [""] * (max_len - len(r)) for r in rows]

            header = [clean_text(h) for h in rows[0]]
            data = rows[1:]

            final_header = []
            seen = {}
            for i, h in enumerate(header):
                name = h if h else f"Spalte_{i + 1}"
                if name in seen:
                    seen[name] += 1
                    name = f"{name}_{seen[name]}"
                else:
                    seen[name] = 1
                final_header.append(name)

            df = pd.DataFrame(data, columns=final_header)
            df = df.replace("", pd.NA).dropna(how="all").fillna("")
            result[sheet_name] = df

    return result


def find_column(columns: List[str], options: List[str]) -> str:
    normalized = {clean_text(c).lower(): c for c in columns}
    for opt in options:
        if opt.lower() in normalized:
            return normalized[opt.lower()]
    return ""


def excel_to_soll_lieferungen(sheets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    records = []

    for sheet_name, df in sheets.items():
        if df.empty:
            continue

        columns = list(df.columns)

        col_csb = find_column(columns, ["CSB", "CSB Nummer", "CSB-Nr", "CSB Nr"])
        col_sap = find_column(columns, ["SAP", "SAP Nummer", "SAP-Nr", "SAP Nr"])
        col_name = find_column(columns, ["Name", "Kunde", "Kundenname"])
        col_strasse = find_column(columns, ["Strasse", "Straße"])
        col_plz = find_column(columns, ["Plz", "PLZ"])
        col_ort = find_column(columns, ["Ort"])

        day_cols = {}
        for day in TAG_SPALTEN:
            found = find_column(columns, [day])
            if found:
                day_cols[day] = found

        if not col_csb or not day_cols:
            continue

        for excel_index, row in df.iterrows():
            csb = norm_num(row.get(col_csb, ""))
            if not csb:
                continue

            base = {
                "Quelle": "EXCEL",
                "Quelle_Blatt": sheet_name,
                "Excel_Zeile": int(excel_index) + 2,
                "CSB": csb,
                "SAP": norm_num(row.get(col_sap, "")) if col_sap else "",
                "Kunde": clean_text(row.get(col_name, "")) if col_name else "",
                "Strasse": clean_text(row.get(col_strasse, "")) if col_strasse else "",
                "PLZ": norm_num(row.get(col_plz, "")) if col_plz else "",
                "Ort": clean_text(row.get(col_ort, "")) if col_ort else "",
            }

            for day, col in day_cols.items():
                tour = norm_tour(row.get(col, ""))
                if not tour:
                    continue

                rec = dict(base)
                rec["Liefertag"] = day
                rec["Soll_Tour_Excel"] = tour
                records.append(rec)

    out = pd.DataFrame(records)
    if out.empty:
        return pd.DataFrame(
            columns=["Quelle", "Quelle_Blatt", "Excel_Zeile", "CSB", "SAP", "Kunde", "Strasse", "PLZ", "Ort", "Liefertag", "Soll_Tour_Excel"]
        )

    out = out.sort_values(["Quelle_Blatt", "CSB", "Liefertag", "Soll_Tour_Excel"], kind="stable").reset_index(drop=True)
    return out


# -------------------------------------------------------------------
# Abgleich
# -------------------------------------------------------------------

def build_lookup(df: pd.DataFrame, tour_col: str) -> Dict[Tuple[str, str], List[dict]]:
    lookup: Dict[Tuple[str, str], List[dict]] = {}
    if df.empty:
        return lookup

    for _, row in df.iterrows():
        key = (norm_num(row.get("CSB", "")), clean_text(row.get("Liefertag", "")))
        if not key[0] or not key[1]:
            continue
        rec = row.to_dict()
        rec["_tour"] = norm_tour(row.get(tour_col, ""))
        lookup.setdefault(key, []).append(rec)

    return lookup


def compare_excel_vs_csb(excel_soll_df: pd.DataFrame, csb_df: pd.DataFrame):
    csb_basis = csb_df.copy()
    if csb_basis.empty:
        csb_basis = pd.DataFrame(columns=["CSB", "Liefertag", "Tour", "Kunde", "Strasse", "PLZ", "Ort"])
    else:
        csb_basis["CSB"] = csb_basis["CSB"].map(norm_num)
        csb_basis["Tour"] = csb_basis["Tour"].map(norm_tour)

    excel_lookup = build_lookup(excel_soll_df, "Soll_Tour_Excel")
    csb_lookup = build_lookup(csb_basis, "Tour")

    abw_rows = []
    ok_rows = []

    # Excel ist maßgeblich: alles aus Excel muss so in CSB stehen.
    for _, xrow in excel_soll_df.iterrows():
        csb = norm_num(xrow.get("CSB", ""))
        day = clean_text(xrow.get("Liefertag", ""))
        soll_tour = norm_tour(xrow.get("Soll_Tour_Excel", ""))
        key = (csb, day)
        actual_rows = csb_lookup.get(key, [])
        actual_tours = sorted({norm_tour(r.get("Tour", "")) or r.get("_tour", "") for r in actual_rows if (norm_tour(r.get("Tour", "")) or r.get("_tour", ""))})

        base = {
            "Status": "",
            "Korrektur_fuer_CSB": "",
            "CSB": csb,
            "SAP": xrow.get("SAP", ""),
            "Kunde_Excel": xrow.get("Kunde", ""),
            "Strasse_Excel": xrow.get("Strasse", ""),
            "PLZ_Excel": xrow.get("PLZ", ""),
            "Ort_Excel": xrow.get("Ort", ""),
            "Liefertag": day,
            "Soll_Tour_Excel": soll_tour,
            "Ist_Tour_CSB": ", ".join(actual_tours),
            "Quelle_Blatt": xrow.get("Quelle_Blatt", ""),
            "Excel_Zeile": xrow.get("Excel_Zeile", ""),
        }

        if not actual_rows:
            row = dict(base)
            row["Status"] = "FEHLT_IN_CSB"
            row["Korrektur_fuer_CSB"] = f"In CSB anlegen/aktivieren: {day} auf Tour {soll_tour}"
            abw_rows.append(row)
        elif soll_tour not in actual_tours:
            row = dict(base)
            row["Status"] = "TOUR_ABWEICHEND"
            row["Korrektur_fuer_CSB"] = f"In CSB ändern: {day} von Tour {', '.join(actual_tours)} auf Tour {soll_tour}"
            abw_rows.append(row)
        else:
            row = dict(base)
            row["Status"] = "OK"
            row["Korrektur_fuer_CSB"] = "Keine Änderung"
            ok_rows.append(row)

    # Alles in CSB, was nicht in Excel steht, muss aus CSB raus.
    for _, crow in csb_basis.iterrows():
        csb = norm_num(crow.get("CSB", ""))
        day = clean_text(crow.get("Liefertag", ""))
        ist_tour = norm_tour(crow.get("Tour", ""))
        key = (csb, day)

        expected_rows = excel_lookup.get(key, [])
        expected_tours = sorted({norm_tour(r.get("Soll_Tour_Excel", "")) or r.get("_tour", "") for r in expected_rows})

        if not expected_rows:
            abw_rows.append(
                {
                    "Status": "ZU_VIEL_IN_CSB",
                    "Korrektur_fuer_CSB": f"In CSB entfernen/deaktivieren: {day} auf Tour {ist_tour}",
                    "CSB": csb,
                    "SAP": "",
                    "Kunde_Excel": "",
                    "Strasse_Excel": "",
                    "PLZ_Excel": "",
                    "Ort_Excel": "",
                    "Liefertag": day,
                    "Soll_Tour_Excel": "",
                    "Ist_Tour_CSB": ist_tour,
                    "Quelle_Blatt": "",
                    "Excel_Zeile": "",
                    "Kunde_CSB": crow.get("Kunde", ""),
                    "Strasse_CSB": crow.get("Strasse", ""),
                    "PLZ_CSB": crow.get("PLZ", ""),
                    "Ort_CSB": crow.get("Ort", ""),
                }
            )
        elif ist_tour not in expected_tours:
            # Wird schon als TOUR_ABWEICHEND aus Excel-Sicht aufgeführt.
            pass

    abweichungen_df = pd.DataFrame(abw_rows)
    ok_df = pd.DataFrame(ok_rows)

    if not abweichungen_df.empty:
        cols = [
            "Status", "Korrektur_fuer_CSB", "CSB", "SAP",
            "Kunde_Excel", "Kunde_CSB", "Strasse_Excel", "Strasse_CSB",
            "PLZ_Excel", "PLZ_CSB", "Ort_Excel", "Ort_CSB",
            "Liefertag", "Soll_Tour_Excel", "Ist_Tour_CSB", "Quelle_Blatt", "Excel_Zeile",
        ]
        for col in cols:
            if col not in abweichungen_df.columns:
                abweichungen_df[col] = ""
        abweichungen_df = abweichungen_df[cols].sort_values(["Status", "CSB", "Liefertag"], kind="stable").reset_index(drop=True)

    excel_csbs = set(excel_soll_df["CSB"].map(norm_num)) if not excel_soll_df.empty else set()
    csb_csbs = set(csb_basis["CSB"].map(norm_num)) if not csb_basis.empty else set()

    kunden_fehlen_komplett = excel_soll_df[~excel_soll_df["CSB"].map(norm_num).isin(csb_csbs)].copy() if not excel_soll_df.empty else pd.DataFrame()
    kunden_zu_viel_komplett = csb_basis[~csb_basis["CSB"].map(norm_num).isin(excel_csbs)].copy() if not csb_basis.empty else pd.DataFrame()

    return abweichungen_df, ok_df, kunden_fehlen_komplett, kunden_zu_viel_komplett


def make_result_excel(
    excel_basis_df: pd.DataFrame,
    csb_basis_df: pd.DataFrame,
    abweichungen_df: pd.DataFrame,
    ok_df: pd.DataFrame,
    kunden_fehlen_komplett_df: pd.DataFrame,
    kunden_zu_viel_komplett_df: pd.DataFrame,
    csb_touren_df: pd.DataFrame,
) -> bytes:
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        abweichungen_df.to_excel(writer, sheet_name="Abweichungen_CSB_anpassen", index=False)
        excel_basis_df.to_excel(writer, sheet_name="Excel_Basis_massgeblich", index=False)
        csb_basis_df.to_excel(writer, sheet_name="CSB_Basis_TXT", index=False)
        ok_df.to_excel(writer, sheet_name="OK", index=False)
        kunden_fehlen_komplett_df.to_excel(writer, sheet_name="Kunden_fehlen_in_CSB", index=False)
        kunden_zu_viel_komplett_df.to_excel(writer, sheet_name="Kunden_zu_viel_in_CSB", index=False)
        csb_touren_df.to_excel(writer, sheet_name="CSB_Touren_Pruefung", index=False)

        wb = writer.book
        from copy import copy
        from openpyxl.styles import Font, PatternFill, Alignment

        header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font = copy(header_font)
                cell.fill = copy(header_fill)
                cell.alignment = copy(header_alignment)
            for col in ws.columns:
                letter = col[0].column_letter
                max_len = 0
                for cell in col:
                    value = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(value))
                ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 55)

    return output.getvalue()


# -------------------------------------------------------------------
# Oberfläche
# -------------------------------------------------------------------

col_u1, col_u2 = st.columns(2)

with col_u1:
    uploaded_txt = st.file_uploader("1. CSB Ladeplan als TXT hochladen", type=["txt"])

with col_u2:
    uploaded_excel = st.file_uploader("2. Maßgebliche Excel hochladen", type=["xlsx"])

if uploaded_txt is None or uploaded_excel is None:
    st.info("Bitte beide Dateien hochladen: CSB-TXT und Excel-Vorgabe.")
    st.stop()

try:
    with st.spinner("Dateien werden gelesen und verglichen..."):
        txt_text = decode_txt_bytes(uploaded_txt.getvalue())
        csb_kunden_df, csb_touren_df, csb_pruefung_df = parse_csb_ladeplan(txt_text)

        excel_sheets = read_excel_first_two_sheets(uploaded_excel.getvalue())
        excel_basis_df = excel_to_soll_lieferungen(excel_sheets)

        abweichungen_df, ok_df, kunden_fehlen_komplett_df, kunden_zu_viel_komplett_df = compare_excel_vs_csb(
            excel_basis_df,
            csb_kunden_df,
        )

    if csb_kunden_df.empty:
        st.error("Im CSB-TXT wurden keine Kunden erkannt.")
        st.stop()

    if excel_basis_df.empty:
        st.error("In den ersten beiden Excel-Blättern wurden keine verwertbaren Liefertage gefunden.")
        st.write("Erwartete Spalten: CSB, SAP, Name, Strasse, Plz, Ort, Mo, Die, Mitt, Don, Fr, Sam")
        st.stop()

    status_counts = abweichungen_df["Status"].value_counts().to_dict() if not abweichungen_df.empty else {}

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Excel-Lieferungen", f"{len(excel_basis_df):,}".replace(",", "."))
    c2.metric("CSB-Lieferungen", f"{len(csb_kunden_df):,}".replace(",", "."))
    c3.metric("Abweichungen", f"{len(abweichungen_df):,}".replace(",", "."))
    c4.metric("Fehlt in CSB", f"{status_counts.get('FEHLT_IN_CSB', 0):,}".replace(",", "."))
    c5.metric("Zu viel in CSB", f"{status_counts.get('ZU_VIEL_IN_CSB', 0):,}".replace(",", "."))

    tour_abw = status_counts.get("TOUR_ABWEICHEND", 0)
    if len(abweichungen_df) == 0:
        st.success("Keine Abweichungen gefunden. CSB passt zur Excel-Vorgabe.")
    else:
        st.warning(
            f"Es wurden {len(abweichungen_df)} Abweichungen gefunden. "
            f"Davon {tour_abw} mit falscher Tour."
        )

    result_excel = make_result_excel(
        excel_basis_df,
        csb_kunden_df,
        abweichungen_df,
        ok_df,
        kunden_fehlen_komplett_df,
        kunden_zu_viel_komplett_df,
        csb_touren_df,
    )

    result_csv = abweichungen_df.to_csv(index=False, sep=";", encoding="utf-8-sig").encode("utf-8-sig")

    d1, d2 = st.columns(2)
    d1.download_button(
        "Abgleich als Excel herunterladen",
        data=result_excel,
        file_name="CSB_Abgleich_Excel_massgeblich.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    d2.download_button(
        "Abweichungen als CSV herunterladen",
        data=result_csv,
        file_name="CSB_Abweichungen.csv",
        mime="text/csv",
        use_container_width=True,
    )

    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs(
        [
            "Abweichungen",
            "Excel Basis",
            "CSB Basis",
            "OK",
            "Kunden fehlen komplett",
            "Kunden zu viel komplett",
        ]
    )

    with tab1:
        st.dataframe(abweichungen_df, use_container_width=True, hide_index=True)

    with tab2:
        st.caption("Maßgeblicher Stand aus den ersten beiden Excel-Blättern.")
        st.dataframe(excel_basis_df, use_container_width=True, hide_index=True)

    with tab3:
        st.caption("Aus dem CSB-TXT erkannter Stand.")
        st.dataframe(csb_kunden_df, use_container_width=True, hide_index=True)

    with tab4:
        st.dataframe(ok_df, use_container_width=True, hide_index=True)

    with tab5:
        st.dataframe(kunden_fehlen_komplett_df, use_container_width=True, hide_index=True)

    with tab6:
        st.dataframe(kunden_zu_viel_komplett_df, use_container_width=True, hide_index=True)

except Exception:
    st.error("Fehler beim Verarbeiten der Dateien.")
    st.code(traceback.format_exc())
